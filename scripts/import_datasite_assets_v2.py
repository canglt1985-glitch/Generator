"""
Import dữ liệu tài sản hạ tầng v2:
- Tải sản tổ accu từ "to accu.xlsx" lồng vào tủ nguồn cha tương ứng.
- Tải sản accu đề và ATS từ "datasite.xlsx" lồng vào máy phát điện cha tương ứng.
- Giữ các tài sản khác (Máy lạnh, CWDM, NLMT) bình thường.
- Cập nhật infrastructure_info của datasites trong Supabase V2.
"""
import os, json, re
from collections import defaultdict
import datetime
import openpyxl
from supabase import create_client

# === CONFIG ===
ORIGINAL_EXCEL_PATH = "/Users/cang_it/Library/CloudStorage/GoogleDrive-canglt1985@gmail.com/My Drive/datasite/datasite.xlsx"
NEW_ACCU_EXCEL_PATH = "/Users/cang_it/Library/CloudStorage/GoogleDrive-canglt1985@gmail.com/My Drive/datasite/to accu.xlsx"

SUPABASE_URL = os.environ.get("SUPABASE_URL_V2", "https://lnmoczxjweuifacqujcu.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY_V2")

if not SUPABASE_KEY:
    # Đọc từ file .env
    env_path = os.path.join(os.path.dirname(__file__), '..', 'tvt3_v2', '.env')
    if os.path.exists(env_path):
        for line in open(env_path):
            if line.startswith('VITE_SUPABASE_ANON_KEY='):
                SUPABASE_KEY = line.split('=', 1)[1].strip()
            if line.startswith('VITE_SUPABASE_URL='):
                SUPABASE_URL = line.split('=', 1)[1].strip()

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# === HELPERS ===
def clean_val(v):
    """Clean value and handle dates correctly."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    if s == '' or s.upper() == 'NONE' or s.upper() == 'KHÔNG CÓ':
        return None
    return s

def read_sheet(wb, sheet_name):
    """Đọc sheet thành list of dicts."""
    ws = wb[sheet_name]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) == 0 or row[0] is None:
            continue
        d = {}
        for i, h in enumerate(headers):
            if i < len(row):
                d[h] = row[i]
            else:
                d[h] = None
        rows.append(d)
    return rows

def extract_number(text):
    """Trích xuất số trong ngoặc đơn (ví dụ: 'MÁY PHÁT ĐIỆN (1)' -> 1)."""
    if not text:
        return None
    match = re.search(r'\((\d+)\)', str(text))
    if match:
        return int(match.group(1))
    return None

def main():
    print("📖 Đang đọc file datasite.xlsx gốc...")
    wb_orig = openpyxl.load_workbook(ORIGINAL_EXCEL_PATH, data_only=True)
    mpd_rows = read_sheet(wb_orig, 'MPD')
    accu_de_rows = read_sheet(wb_orig, 'Accu de')
    ats_rows = read_sheet(wb_orig, 'ATS')
    maylanh_rows = read_sheet(wb_orig, 'MayLanh')
    tunguon_rows = read_sheet(wb_orig, 'tu nguon')
    cwdm_rows = read_sheet(wb_orig, 'CWDM')
    nlmt_rows = read_sheet(wb_orig, 'NLMT')
    wb_orig.close()

    print("📖 Đang đọc file to accu.xlsx mới...")
    wb_new_accu = openpyxl.load_workbook(NEW_ACCU_EXCEL_PATH, data_only=True)
    toaccu_rows = read_sheet(wb_new_accu, 'DataSite')
    wb_new_accu.close()

    print(f"  MPĐ: {len(mpd_rows)} | Accu đề: {len(accu_de_rows)} | ATS: {len(ats_rows)}")
    print(f"  Máy lạnh: {len(maylanh_rows)} | Tủ nguồn: {len(tunguon_rows)} | Tổ accu mới: {len(toaccu_rows)}")
    print(f"  CWDM: {len(cwdm_rows)} | NLMT: {len(nlmt_rows)}")

    # === LOGIC GHÉP NỐI CHA - CON ===
    
    # 1. Ghép Accu đề và ATS vào Máy phát điện cha
    # Khởi tạo danh sách máy phát điện theo site
    mpd_by_site = defaultdict(list)
    for r in mpd_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        mpd_by_site[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Máy phát điện')),
            "cong_suat": clean_val(r.get('Công suất Máy phát điện')),
            "nhien_lieu": clean_val(r.get('Nhiên liệu')),
            "serial": clean_val(r.get('Serial máy phát điện')),
            "product_code": clean_val(r.get('Product_code máy phát điện')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            "bao_duong": clean_val(r.get('Thời hạn bảo dưỡng')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "accu_de": [],
            "ats": []
        })

    # Ghép accu đề vào máy phát điện
    for r in accu_de_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        
        accu_item = {
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Accu đề')),
            "loai": clean_val(r.get('Loại accu đề')),
            "serial": clean_val(r.get('Serial accu đề')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng (theo hợp đồng trang bị)')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
        }
        
        mpds = mpd_by_site[sid]
        if not mpds:
            # Tạo máy phát điện giả định nếu không tìm thấy
            mpd_by_site[sid].append({
                "ten": "MÁY PHÁT ĐIỆN (1)",
                "accu_de": [accu_item],
                "ats": []
            })
            continue
            
        # Tìm máy phát điện khớp số thứ tự
        accu_num = extract_number(accu_item["ten"])
        matched = False
        if accu_num is not None:
            for mpd in mpds:
                mpd_num = extract_number(mpd["ten"])
                if mpd_num == accu_num:
                    mpd["accu_de"].append(accu_item)
                    matched = True
                    break
        
        if not matched:
            # Mặc định gắn vào máy phát điện đầu tiên
            mpds[0]["accu_de"].append(accu_item)

    # Ghép ATS vào máy phát điện
    for r in ats_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        
        ats_item = {
            "ten": clean_val(r.get('Tên đối tượng')) or 'ATS',
            "nhan_hieu": clean_val(r.get('Nhãn hiệu ATS')),
            "serial": clean_val(r.get('Serial ATS')),
            "product_code": clean_val(r.get('Product_code ATS')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng (theo hợp đồng trang bị)')),
        }
        
        mpds = mpd_by_site[sid]
        if not mpds:
            mpd_by_site[sid].append({
                "ten": "MÁY PHÁT ĐIỆN (1)",
                "accu_de": [],
                "ats": [ats_item]
            })
            continue
            
        ats_num = extract_number(ats_item["ten"])
        matched = False
        if ats_num is not None:
            for mpd in mpds:
                mpd_num = extract_number(mpd["ten"])
                if mpd_num == ats_num:
                    mpd["ats"].append(ats_item)
                    matched = True
                    break
                    
        if not matched:
            mpds[0]["ats"].append(ats_item)

    # 2. Ghép Tổ accu mới vào Tủ nguồn DC cha
    tunguon_by_site = defaultdict(list)
    for r in tunguon_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        tunguon_by_site[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Tủ nguồn (Vender)')),
            "so_khe_rectifier": clean_val(r.get('Số lượng khe rectifier')),
            "so_luong_rectifier": clean_val(r.get('Số lượng rectifier')),
            "cong_suat_rectifier": clean_val(r.get('Công suất Rectifier (W)')),
            "thoi_gian_backup": clean_val(r.get('Thời gian Backup (phút)')),
            "serial": clean_val(r.get('Serial tủ nguồn')),
            "product_code": clean_val(r.get('Product_code tủ nguồn (model/part name)')),
            "product_code_rectifier": clean_val(r.get('Product code rectifier')),
            "dong_tai": clean_val(r.get('Dòng tải thiết bị (A)')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            "to_accu": []
        })

    # Ghép tổ accu vào tủ nguồn dựa trên cột "Tên đối tượng cha"
    for r in toaccu_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
            
        parent_name = clean_val(r.get('Tên đối tượng cha')) or 'TỦ NGUỒN (1)'
        
        accu_item = {
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Accu')),
            "loai": clean_val(r.get('Loại Accu')),
            "dung_luong": clean_val(r.get('Dung lượng bình Accu')),
            "so_luong_binh": clean_val(r.get('Số lượng Bình Accu')),
            "serial": clean_val(r.get('Serial ACCU')),
            "product_code": clean_val(r.get('Product_code ACCU')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng (theo hợp đồng trang bị)')),
        }
        
        tns = tunguon_by_site[sid]
        
        # Tìm tủ nguồn có tên trùng với Tên đối tượng cha
        matched = False
        for tn in tns:
            if tn["ten"] == parent_name:
                tn["to_accu"].append(accu_item)
                matched = True
                break
                
        if not matched:
            # Tạo tủ nguồn mới nếu trạm chưa có tủ nguồn này
            new_tn = {
                "ten": parent_name,
                "nhan_hieu": None,
                "to_accu": [accu_item]
            }
            tunguon_by_site[sid].append(new_tn)

    # 3. Gom Máy lạnh
    maylanh_by_site = defaultdict(list)
    for r in maylanh_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        maylanh_by_site[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Máy lạnh')),
            "cong_suat": clean_val(r.get('Công suất lạnh (BTU)')),
            "loai": clean_val(r.get('Loại máy lạnh')),
            "serial": clean_val(r.get('Serial máy lạnh (serial dàn lạnh/serial dàn nóng)')),
            "product_code": clean_val(r.get('Product_code máy lạnh')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
        })

    # 4. Gom CWDM
    cwdm_by_site = defaultdict(list)
    for r in cwdm_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        cwdm_by_site[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "ten_thiet_bi": clean_val(r.get('Tên thiết bị CWDM')),
            "loai": clean_val(r.get('Loại thiết bị CWDM')),
            "ma_thiet_bi": clean_val(r.get('Mã thiết bị CWDM')),
            "hang_sx": clean_val(r.get('Hãng sản xuất CWDM')),
            "serial": clean_val(r.get('Serial bộ thiết bị CWDM')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "ghi_chu": clean_val(r.get('Ghi chú')),
        })

    # 5. Gom NLMT
    nlmt_by_site = {}
    for r in nlmt_rows:
        sid = clean_val(r.get('SITE_ID'))
        if not sid:
            continue
        nlmt_by_site[sid] = {
            "cong_suat": clean_val(r.get('Công suất hệ thống NLMT')),
            "loai_he_thong": clean_val(r.get('Loại hệ thống NLMT')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "inverter": {
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Inverter')),
                "product_code": clean_val(r.get('Product_code Inverter')),
                "cong_suat": clean_val(r.get('Công suất Inverter')),
                "serial": clean_val(r.get('Serial Inverter')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành Inverter')),
            },
            "tam_pin": {
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Pin NLMT')),
                "product_code": clean_val(r.get('Product_code Pin NLMT')),
                "cong_suat": clean_val(r.get('Công suất Pin NLMT')),
                "so_luong": clean_val(r.get('Số lượng Pin NLMT')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành tấm pin')),
            },
            "sim_giam_sat": clean_val(r.get('Số thuê bao SIM giám sát')),
        }

    # Tổng hợp danh sách site_ids
    all_sids = set()
    all_sids.update(mpd_by_site.keys())
    all_sids.update(tunguon_by_site.keys())
    all_sids.update(maylanh_by_site.keys())
    all_sids.update(cwdm_by_site.keys())
    all_sids.update(nlmt_by_site.keys())

    print(f"\n🔧 Đang tổng hợp dữ liệu cho {len(all_sids)} trạm...")

    # Cập nhật DB
    success = 0
    errors = 0
    for sid in sorted(all_sids):
        infra = {}
        
        # MPĐ lồng accu đề & ATS
        if sid in mpd_by_site and mpd_by_site[sid]:
            infra["may_phat_dien"] = {
                "mpd": mpd_by_site[sid]
            }
            
        # Tủ nguồn lồng tổ accu
        if sid in tunguon_by_site and tunguon_by_site[sid]:
            infra["nguon_dien"] = {
                "tu_nguon": tunguon_by_site[sid]
            }

        # Máy lạnh
        if sid in maylanh_by_site and maylanh_by_site[sid]:
            infra["may_lanh"] = maylanh_by_site[sid]

        # CWDM
        if sid in cwdm_by_site and cwdm_by_site[sid]:
            infra["cwdm"] = cwdm_by_site[sid]

        # NLMT
        if sid in nlmt_by_site:
            infra["nang_luong_mat_troi"] = nlmt_by_site[sid]

        try:
            resp = supabase.table('datasites').update({
                'infrastructure_info': infra
            }).eq('site_id_old', sid).execute()
            
            if resp.data and len(resp.data) > 0:
                success += 1
            else:
                print(f"  ⚠️ Trạm {sid} không tồn tại trên DB (site_id_old)")
                errors += 1
        except Exception as e:
            print(f"  ❌ Lỗi khi cập nhật trạm {sid}: {e}")
            errors += 1

    print(f"\n✅ Hoàn tất nạp dữ liệu!")
    print(f"  Thành công: {success} trạm")
    print(f"  Lỗi/Không tìm thấy: {errors} trạm")

    # Sample verification
    print(f"\n🔍 Kiểm tra mẫu (DNCM00)...")
    sample = supabase.table('datasites').select('site_id, site_id_old, infrastructure_info').eq('site_id_old', 'DNCM00').execute()
    if sample.data:
        infra = sample.data[0].get('infrastructure_info', {})
        mpd_list = infra.get('may_phat_dien', {}).get('mpd', [])
        print(f"  MPĐ chính: {len(mpd_list)} máy")
        for idx, m in enumerate(mpd_list):
            print(f"    - Máy phát {idx+1}: {m.get('ten')} (Accu đề con: {len(m.get('accu_de', []))}, ATS con: {len(m.get('ats', []))})")
            
        tn_list = infra.get('nguon_dien', {}).get('tu_nguon', [])
        print(f"  Tủ nguồn DC: {len(tn_list)} tủ")
        for idx, t in enumerate(tn_list):
            print(f"    - Tủ nguồn {idx+1}: {t.get('ten')} (Tổ accu con: {len(t.get('to_accu', []))})")

if __name__ == '__main__':
    main()
