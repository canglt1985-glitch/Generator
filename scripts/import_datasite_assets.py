"""
Import dữ liệu tài sản hạ tầng từ datasite.xlsx vào cột infrastructure_info của bảng datasites (Supabase V2).

Nhóm gom:
  - MPĐ + Accu đề + ATS → may_phat_dien (gom theo site_id)
  - Tủ nguồn + Tổ accu → tu_nguon (gom theo site_id)
  - Máy lạnh → may_lanh
  - CWDM → cwdm
  - NLMT → nang_luong_mat_troi
"""
import os, json
from collections import defaultdict
import openpyxl
from supabase import create_client

# === CONFIG ===
EXCEL_PATH = "/Users/cang_it/Library/CloudStorage/GoogleDrive-canglt1985@gmail.com/My Drive/datasite/datasite.xlsx"

SUPABASE_URL = os.environ.get("SUPABASE_URL_V2", "https://lnmoczxjweuifacqujcu.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY_V2")

if not SUPABASE_KEY:
    # Thử đọc từ .env của tvt3_v2
    env_path = os.path.join(os.path.dirname(__file__), '..', 'tvt3_v2', '.env')
    if os.path.exists(env_path):
        for line in open(env_path):
            if line.startswith('VITE_SUPABASE_ANON_KEY='):
                SUPABASE_KEY = line.split('=', 1)[1].strip()
            if line.startswith('VITE_SUPABASE_URL='):
                SUPABASE_URL = line.split('=', 1)[1].strip()

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# === HELPERS ===
def read_sheet(wb, sheet_name):
    """Đọc sheet thành list of dicts với header row 1."""
    ws = wb[sheet_name]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) == 0 or not row[0]:
            continue
        d = {}
        for i, h in enumerate(headers):
            if i < len(row) and row[i] is not None:
                val = row[i]
                d[h] = str(val).strip() if not isinstance(val, (int, float)) else val
            else:
                d[h] = ''
        rows.append(d)
    return rows

def clean_val(v):
    """Clean empty/None values."""
    if v is None or str(v).strip() == '' or str(v).strip().upper() == 'NONE':
        return None
    return str(v).strip() if not isinstance(v, (int, float)) else v


def build_mpd_group(mpd_rows, accu_rows, ats_rows):
    """Gom MPĐ + Accu đề + ATS theo site_id."""
    # Index accu đề và ATS theo site_id
    accu_by_site = defaultdict(list)
    for r in accu_rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if sid:
            accu_by_site[sid].append({
                "ten": clean_val(r.get('Tên đối tượng')),
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Accu đề')),
                "loai": clean_val(r.get('Loại Accu đề')),
                "product_code": clean_val(r.get('Product_code Accu đề')),
                "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
                "tinh_trang": clean_val(r.get('Trạng thái')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            })
    
    ats_by_site = defaultdict(list)
    for r in ats_rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if sid:
            ats_by_site[sid].append({
                "nhan_hieu": clean_val(r.get('Nhãn hiệu ATS')),
                "serial": clean_val(r.get('Serial ATS')),
                "product_code": clean_val(r.get('Product_code ATS')),
                "tinh_trang": clean_val(r.get('Trạng thái')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
                "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng (theo hợp đồng trang bị)')),
            })
    
    # Gom MPĐ chính
    result = defaultdict(list)
    for r in mpd_rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if not sid:
            continue
        mpd = {
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Máy phát điện')),
            "cong_suat": clean_val(r.get('Công suất Máy phát điện')),
            "nhien_lieu": clean_val(r.get('Nhiên liệu')),
            "serial": clean_val(r.get('Serial máy phát điện')),
            "product_code": clean_val(r.get('Product_code máy phát điện')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
            "bao_duong": clean_val(r.get('Thời hạn bảo dưỡng')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
        }
        result[sid].append(mpd)
    
    # Gắn accu đề và ATS vào nhóm MPĐ
    all_sids = set(result.keys()) | set(accu_by_site.keys()) | set(ats_by_site.keys())
    final = {}
    for sid in all_sids:
        final[sid] = {
            "mpd": result.get(sid, []),
            "accu_de": accu_by_site.get(sid, []),
            "ats": ats_by_site.get(sid, []),
        }
    return final


def build_tunguon_group(tunguon_rows, toaccu_rows):
    """Gom Tủ nguồn + Tổ accu theo site_id."""
    toaccu_by_site = defaultdict(list)
    for r in toaccu_rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if sid:
            toaccu_by_site[sid].append({
                "ten": clean_val(r.get('Tên đối tượng')),
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Accu')),
                "loai": clean_val(r.get('Loại Accu')),
                "dung_luong": clean_val(r.get('Dung lượng bình Accu')),
                "so_luong_binh": clean_val(r.get('Số lượng Bình Accu')),
                "serial": clean_val(r.get('Serial ACCU')),
                "product_code": clean_val(r.get('Product_code ACCU')),
                "ma_tai_san": clean_val(r.get('Mã tài sản')),
                "tinh_trang": clean_val(r.get('Trạng thái')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
                "bao_duong": clean_val(r.get('Thời hạn bảo dưỡng')),
                "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            })
    
    result = defaultdict(list)
    for r in tunguon_rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if not sid:
            continue
        tn = {
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Tủ nguồn (Vender)')),
            "so_khe_rectifier": clean_val(r.get('Số lượng khe rectifier')),
            "so_luong_rectifier": clean_val(r.get('Số lượng rectifier')),
            "cong_suat_rectifier": clean_val(r.get('Công suất Rectifier (W)')),
            "thoi_gian_backup": clean_val(r.get('Thời gian Backup (phút)')),
            "serial": clean_val(r.get('Serial tủ nguồn')),
            "product_code": clean_val(r.get('Product_code tủ nguồn (model/part name)')),
            "product_code_rectifier": clean_val(r.get('Product code rectifier')),
            "dong_tai": clean_val(r.get('Dòng tải thiết bị (A)')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
        }
        result[sid].append(tn)
    
    all_sids = set(result.keys()) | set(toaccu_by_site.keys())
    final = {}
    for sid in all_sids:
        final[sid] = {
            "tu_nguon": result.get(sid, []),
            "to_accu": toaccu_by_site.get(sid, []),
        }
    return final


def build_maylanh(rows):
    result = defaultdict(list)
    for r in rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if not sid:
            continue
        result[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "nhan_hieu": clean_val(r.get('Nhãn hiệu Máy lạnh')),
            "cong_suat": clean_val(r.get('Công suất lạnh (BTU)')),
            "loai": clean_val(r.get('Loại máy lạnh')),
            "serial": clean_val(r.get('Serial máy lạnh (serial dàn lạnh/serial dàn nóng)')),
            "product_code": clean_val(r.get('Product_code máy lạnh')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "bao_hanh": clean_val(r.get('Thời hạn bảo hành')),
        })
    return result


def build_cwdm(rows):
    result = defaultdict(list)
    for r in rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if not sid:
            continue
        result[sid].append({
            "ten": clean_val(r.get('Tên đối tượng')),
            "ten_thiet_bi": clean_val(r.get('Tên thiết bị CWDM')),
            "loai": clean_val(r.get('Loại thiết bị CWDM')),
            "ma_thiet_bi": clean_val(r.get('Mã thiết bị CWDM')),
            "hang_sx": clean_val(r.get('Hãng sản xuất CWDM')),
            "serial": clean_val(r.get('Serial bộ thiết bị CWDM')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "ghi_chu": clean_val(r.get('Ghi chú')),
        })
    return result


def build_nlmt(rows):
    result = {}
    for r in rows:
        sid = str(r.get('SITE_ID', '')).strip()
        if not sid:
            continue
        result[sid] = {
            "cong_suat": clean_val(r.get('Công suất hệ thống NLMT')),
            "loai_he_thong": clean_val(r.get('Loại hệ thống NLMT')),
            "ma_tai_san": clean_val(r.get('Mã tài sản')),
            "ngay_su_dung": clean_val(r.get('Ngày đưa vào sử dụng')),
            "tinh_trang": clean_val(r.get('Trạng thái')),
            "inverter": {
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Inverter')),
                "product_code": clean_val(r.get('Product_code Inverter')),
                "cong_suat": clean_val(r.get('Công suất Inverter')),
                "serial": clean_val(r.get('Serial Inverter')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành Inverter')),
            },
            "tam_pin": {
                "nhan_hieu": clean_val(r.get('Nhãn hiệu Pin NLMT')),
                "product_code": clean_val(r.get('Product_code Pin NLMT')),
                "cong_suat": clean_val(r.get('Công suất Pin NLMT')),
                "so_luong": clean_val(r.get('Số lượng Pin NLMT')),
                "bao_hanh": clean_val(r.get('Thời hạn bảo hành tấm pin')),
            },
            "sim_giam_sat": clean_val(r.get('Số thuê bao SIM giám sát')),
        }
    return result


# === MAIN ===
def main():
    print("📖 Đang đọc file datasite.xlsx...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    
    # Đọc tất cả sheet
    mpd_rows = read_sheet(wb, 'MPD')
    accu_rows = read_sheet(wb, 'Accu de')
    ats_rows = read_sheet(wb, 'ATS')
    maylanh_rows = read_sheet(wb, 'MayLanh')
    tunguon_rows = read_sheet(wb, 'tu nguon')
    toaccu_rows = read_sheet(wb, 'to accu')
    cwdm_rows = read_sheet(wb, 'CWDM')
    nlmt_rows = read_sheet(wb, 'NLMT')
    wb.close()
    
    print(f"  MPĐ: {len(mpd_rows)} | Accu đề: {len(accu_rows)} | ATS: {len(ats_rows)}")
    print(f"  Máy lạnh: {len(maylanh_rows)} | Tủ nguồn: {len(tunguon_rows)} | Tổ accu: {len(toaccu_rows)}")
    print(f"  CWDM: {len(cwdm_rows)} | NLMT: {len(nlmt_rows)}")
    
    # Gom nhóm
    print("\n🔧 Đang gom nhóm dữ liệu...")
    mpd_group = build_mpd_group(mpd_rows, accu_rows, ats_rows)
    tunguon_group = build_tunguon_group(tunguon_rows, toaccu_rows)
    maylanh_data = build_maylanh(maylanh_rows)
    cwdm_data = build_cwdm(cwdm_rows)
    nlmt_data = build_nlmt(nlmt_rows)
    
    # Tổng hợp tất cả site_ids
    all_sids = set()
    all_sids.update(mpd_group.keys())
    all_sids.update(tunguon_group.keys())
    all_sids.update(maylanh_data.keys())
    all_sids.update(cwdm_data.keys())
    all_sids.update(nlmt_data.keys())
    
    print(f"  Tổng: {len(all_sids)} trạm cần cập nhật")
    
    # Build infrastructure_info cho từng trạm
    print("\n🚀 Đang cập nhật database Supabase V2...")
    success = 0
    errors = 0
    
    for sid in sorted(all_sids):
        infra = {}
        
        # Nhóm MPĐ (gồm accu đề + ATS)
        if sid in mpd_group:
            infra["may_phat_dien"] = mpd_group[sid]
        
        # Nhóm Tủ nguồn (gồm tổ accu)
        if sid in tunguon_group:
            infra["nguon_dien"] = tunguon_group[sid]
        
        # Máy lạnh riêng
        if sid in maylanh_data:
            infra["may_lanh"] = maylanh_data[sid]
        
        # CWDM riêng
        if sid in cwdm_data:
            infra["cwdm"] = cwdm_data[sid]
        
        # NLMT riêng
        if sid in nlmt_data:
            infra["nang_luong_mat_troi"] = nlmt_data[sid]
        
        try:
            # Match bằng site_id_old vì Excel dùng mã trạm cũ
            resp = supabase.table('datasites').update({
                'infrastructure_info': infra
            }).eq('site_id_old', sid).execute()
            
            if resp.data and len(resp.data) > 0:
                success += 1
            else:
                print(f"  ⚠️ Site {sid} không tìm thấy trên DB (site_id_old)")
                errors += 1
        except Exception as e:
            print(f"  ❌ Lỗi site {sid}: {e}")
            errors += 1
    
    print(f"\n✅ Hoàn tất!")
    print(f"  Thành công: {success} trạm")
    print(f"  Lỗi/Không tìm thấy: {errors} trạm")
    
    # Sample verification
    print(f"\n🔍 Kiểm tra mẫu (DNCM00)...")
    sample = supabase.table('datasites').select('site_id, site_id_old, infrastructure_info').eq('site_id_old', 'DNCM00').execute()
    if sample.data:
        infra = sample.data[0].get('infrastructure_info', {})
        print(f"  MPĐ: {len(infra.get('may_phat_dien', {}).get('mpd', []))} máy")
        print(f"  Accu đề: {len(infra.get('may_phat_dien', {}).get('accu_de', []))} bộ")
        print(f"  ATS: {len(infra.get('may_phat_dien', {}).get('ats', []))} bộ")
        print(f"  Tủ nguồn: {len(infra.get('nguon_dien', {}).get('tu_nguon', []))} tủ")
        print(f"  Tổ accu: {len(infra.get('nguon_dien', {}).get('to_accu', []))} tổ")
        print(f"  Máy lạnh: {len(infra.get('may_lanh', []))} máy")
        print(f"  CWDM: {len(infra.get('cwdm', []))} bộ")
        has_nlmt = 'nang_luong_mat_troi' in infra
        print(f"  NLMT: {'Có' if has_nlmt else 'Không'}")


if __name__ == '__main__':
    main()
