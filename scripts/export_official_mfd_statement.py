#!/usr/bin/env python3
"""
Official MFD Statement & Invoice Exporter (Mẫu 02A-TTNB_NLMPD & HD)
Xuất hồ sơ thanh toán máy phát điện và hóa đơn nhiên liệu chuẩn đẹp ra file Excel.
"""

import os
import sys
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Supabase setup
try:
    from supabase import create_client
except ImportError:
    pass

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://lnmoczxjweuifacqujcu.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImxubW9jenhqd2V1aWZhY3F1amN1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg2MzcxOTYsImV4cCI6MjA5NDIxMzE5Nn0.C0Si7ChY4T_mxLylSkDNJOUcj9D0uuGW_L4t7p9yONI")

# Special 67 Sites for MobiFone Dong Nai (Group 1)
SPECIAL_67_SITES = {
    'DNCM11', 'DNCM14', 'DNCM15', 'DNCM23', 'DNCM45', 'DNDQ03', 'DNDQ31', 'DNDQ51',
    'DNDQ58', 'DNIDQN1', 'DNLK40', 'DNLK42', 'DNLK27', 'DNLK71', 'DNLK73', 'DNLT20',
    'DNLT29', 'DNLT64', 'DNLT87', 'DNLT91', 'DNLT98', 'DNLTA1', 'DNLTA3', 'DNLTA8',
    'DNLTC8', 'DNLTX5', 'DNNT82SR01', 'DNNTA5', 'DNTN24', 'DNTN43', 'DNTNL2', 'DNTP03',
    'DNTP08', 'DNTP30', 'DNTP42', 'DNTP44', 'DNTP53', 'DNXL37', 'DNXL45', 'DNXL49',
    'DNXL65', 'DNXL75', 'DNXL77', 'DNINTR49', 'DNINTR64', 'DNIAPH29', 'DNIAPH30',
    'DNIAPH35', 'DNIAPH28', 'DNIAPH24', 'DNIAPH12', 'DNIAPH26', 'DNIAPH32', 'DNIBAN03',
    'DNIBAN10', 'DNILTH10'
}

def is_group_1(site_id, site_id_old=""):
    s1 = (site_id or "").upper().strip()
    s2 = (site_id_old or "").upper().strip()
    return s1 in SPECIAL_67_SITES or s2 in SPECIAL_67_SITES

def create_styled_workbook(month=8, year=2026, output_path=None):
    from supabase import create_client
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # 1. Query logs for the month
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year+1}-01-01"
    else:
        end_date = f"{year}-{month+1:02d}-01"
        
    print(f"Fetching logs from {start_date} to {end_date}...")
    logs_res = supabase.from_("generator_logs").select("*").gte("date", start_date).lt("date", end_date).execute()
    logs = logs_res.data or []
    
    invoices_res = supabase.from_("parsed_invoices").select("*").gte("invoice_date", start_date).lt("invoice_date", end_date).execute()
    invoices = invoices_res.data or []
    
    sites_res = supabase.from_("datasites").select("*").execute()
    sites = {s.get("site_id"): s for s in (sites_res.data or [])}

    print(f"Found {len(logs)} generator logs, {len(invoices)} invoices.")

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE')
    )
    thick_bottom = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='medium', color='1F2328')
    )
    double_bottom = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='double', color='1F2328')
    )

    header_fill_blue = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_fill_amber = PatternFill(start_color='B45309', end_color='B45309', fill_type='solid')
    header_fill_green = PatternFill(start_color='047857', end_color='047857', fill_type='solid')
    light_blue = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    light_amber = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    sum_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    def add_02a_sheet(sheet_title, target_logs, group_name):
        ws = wb.create_sheet(title=sheet_title)
        
        ws['A1'] = 'TRUNG TÂM MẠNG LƯỚI MOBIFONE MIỀN NAM'
        ws['A1'].font = Font(name='Arial', size=10, bold=True)
        ws['A2'] = 'ĐƠN VỊ: ĐÀI VIỄN THÔNG ĐỒNG NAI'
        ws['A2'].font = Font(name='Arial', size=10, bold=True)

        ws['A4'] = 'BẢNG KÊ CHI TIẾT NHIÊN LIỆU CHẠY MÁY PHÁT ĐIỆN ĐÃ ĐƯỢC ĐỐI SOÁT'
        ws['A4'].font = Font(name='Arial', size=13, bold=True, color='1E3A8A')
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A4:O4')

        ws['A5'] = f'(Tháng {month:02d}/{year} - {group_name})'
        ws['A5'].font = Font(name='Arial', size=10, italic=True)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A5:O5')

        # Separate Xăng vs Dầu
        xang_list = []
        dau_list = []
        for l in target_logs:
            rd = l.get('run_details') or {}
            st = sites.get(l.get('site_id')) or {}
            nl = (rd.get('nhien_lieu_loai') or rd.get('nhien_lieu') or '').lower()
            lm = (rd.get('loai_may') or st.get('loai_may') or '').lower()
            if 'xăng' in nl or 'xang' in nl or 'kibi' in lm or 'hyundai' in lm:
                xang_list.append((l, rd, st))
            else:
                dau_list.append((l, rd, st))

        tot_m_x = sum(float(r[1].get('thanh_tien') or 0) for r in xang_list)
        tot_m_d = sum(float(r[1].get('thanh_tien') or 0) for r in dau_list)
        tot_m_all = tot_m_x + tot_m_d

        # Top summary
        ws['A7'] = 'STT'; ws['B7'] = 'Nội dung thanh toán'; ws.merge_cells('B7:E7')
        ws['F7'] = 'Số tiền chưa VAT (VNĐ)'; ws.merge_cells('F7:H7')
        ws['I7'] = 'Tiền thuế VAT (VNĐ)'; ws.merge_cells('I7:J7')
        ws['K7'] = 'Tổng tiền có VAT (VNĐ)'; ws.merge_cells('K7:L7')
        ws['M7'] = 'Ghi chú'; ws.merge_cells('M7:O7')
        
        for c in range(1, 16):
            cell = ws.cell(7, c)
            cell.fill = header_fill_blue
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        ws['A8'] = 1; ws['B8'] = f'Nhiên liệu chạy máy phát điện tháng {month:02d}/{year} ({group_name})'
        ws.merge_cells('B8:E8')
        ws['F8'] = round(tot_m_all); ws['F8'].number_format = '#,##0'; ws.merge_cells('F8:H8')
        ws['I8'] = 0; ws['I8'].number_format = '#,##0'; ws.merge_cells('I8:J8')
        ws['K8'] = '=F8+I8'; ws['K8'].number_format = '#,##0'; ws.merge_cells('K8:L8')
        ws['M8'] = 'Theo đối soát thực tế'; ws.merge_cells('M8:O8')
        for c in range(1, 16):
            ws.cell(8, c).font = Font(name='Arial', size=9)
            ws.cell(8, c).border = thin_border
            if c in [6, 9, 11]: ws.cell(8, c).alignment = Alignment(horizontal='right')

        ws['A9'] = 'TỔNG CỘNG THANH TOÁN'; ws.merge_cells('A9:E9')
        ws['F9'] = '=F8'; ws['F9'].number_format = '#,##0'; ws.merge_cells('F9:H9')
        ws['I9'] = '=I8'; ws['I9'].number_format = '#,##0'; ws.merge_cells('I9:J9')
        ws['K9'] = '=K8'; ws['K9'].number_format = '#,##0'; ws.merge_cells('K9:L9')
        for c in range(1, 16):
            ws.cell(9, c).font = Font(name='Arial', size=9, bold=True, color='1E3A8A')
            ws.cell(9, c).fill = sum_fill
            ws.cell(9, c).border = thick_bottom

        # Table headers
        headers_detail = [
            'STT', 'Tên Trạm', 'ID Trạm', 'Công suất máy (kVA)', 'Loại máy nổ',
            'Định mức (L/h)', 'Ngày vận hành', 'Giờ bắt đầu', 'Giờ kết thúc',
            'Thời gian hoạt động (giờ)', 'Nhiên liệu tiêu hao (lít)', 'Đơn giá trước VAT',
            'Thành tiền trước VAT (đồng)', 'Ghi chú', 'Kết quả đối soát'
        ]

        # Section A: Xăng
        ws['A11'] = 'A. MÁY CHẠY XĂNG'; ws['A11'].font = Font(name='Arial', size=11, bold=True, color='1E3A8A')
        for c, h in enumerate(headers_detail, 1):
            cell = ws.cell(12, c, h)
            cell.fill = header_fill_blue
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        cur_row = 13
        stt = 1
        start_xang = cur_row
        for l, rd, st in xang_list:
            ws.cell(cur_row, 1, stt)
            ws.cell(cur_row, 2, l.get('site_id') or '')
            ws.cell(cur_row, 3, st.get('site_id_old') or l.get('site_id') or '')
            ws.cell(cur_row, 4, rd.get('cong_suat_may') or st.get('cong_suat') or 6)
            ws.cell(cur_row, 5, rd.get('loai_may') or st.get('loai_may') or 'KIBI')
            ws.cell(cur_row, 6, float(rd.get('dinh_muc') or 3.44))
            ws.cell(cur_row, 7, l.get('date') or '')
            ws.cell(cur_row, 8, rd.get('gio_bat_dau') or '')
            ws.cell(cur_row, 9, rd.get('gio_ket_thuc') or '')
            ws.cell(cur_row, 10, float(rd.get('thoi_gian_hoat_dong') or 0)).number_format = '0.00'
            ws.cell(cur_row, 11, float(rd.get('nhien_lieu_tieu_hao') or 0)).number_format = '0.00'
            ws.cell(cur_row, 12, float(rd.get('don_gia') or 0)).number_format = '#,##0'
            ws.cell(cur_row, 13, float(rd.get('thanh_tien') or 0)).number_format = '#,##0'
            ws.cell(cur_row, 14, rd.get('ghi_chu') or '')
            ws.cell(cur_row, 15, rd.get('ket_qua_doi_soat') or 'OK')

            for c in range(1, 16):
                ws.cell(cur_row, c).font = Font(name='Arial', size=9)
                ws.cell(cur_row, c).border = thin_border
            stt += 1
            cur_row += 1
        end_xang = cur_row - 1

        # Sum Xang
        ws.cell(cur_row, 1, 'Tổng cộng máy chạy xăng')
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=9)
        ws.cell(cur_row, 10, f'=SUM(J{start_xang}:J{end_xang})').number_format = '0.00'
        ws.cell(cur_row, 11, f'=SUM(K{start_xang}:K{end_xang})').number_format = '0.00'
        ws.cell(cur_row, 13, f'=SUM(M{start_xang}:M{end_xang})').number_format = '#,##0'
        for c in range(1, 16):
            ws.cell(cur_row, c).font = Font(name='Arial', size=9, bold=True, color='1E3A8A')
            ws.cell(cur_row, c).fill = light_blue
            ws.cell(cur_row, c).border = thick_bottom
        sum_xang_row = cur_row
        cur_row += 2

        # Section B: Dầu
        ws.cell(cur_row, 1, 'B. MÁY CHẠY DẦU').font = Font(name='Arial', size=11, bold=True, color='B45309')
        cur_row += 1
        for c, h in enumerate(headers_detail, 1):
            cell = ws.cell(cur_row, c, h)
            cell.fill = header_fill_amber
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        cur_row += 1

        start_dau = cur_row
        stt = 1
        for l, rd, st in dau_list:
            ws.cell(cur_row, 1, stt)
            ws.cell(cur_row, 2, l.get('site_id') or '')
            ws.cell(cur_row, 3, st.get('site_id_old') or l.get('site_id') or '')
            ws.cell(cur_row, 4, rd.get('cong_suat_may') or st.get('cong_suat') or 8.5)
            ws.cell(cur_row, 5, rd.get('loai_may') or st.get('loai_may') or 'HỮU TOÀN')
            ws.cell(cur_row, 6, float(rd.get('dinh_muc') or 3.05))
            ws.cell(cur_row, 7, l.get('date') or '')
            ws.cell(cur_row, 8, rd.get('gio_bat_dau') or '')
            ws.cell(cur_row, 9, rd.get('gio_ket_thuc') or '')
            ws.cell(cur_row, 10, float(rd.get('thoi_gian_hoat_dong') or 0)).number_format = '0.00'
            ws.cell(cur_row, 11, float(rd.get('nhien_lieu_tieu_hao') or 0)).number_format = '0.00'
            ws.cell(cur_row, 12, float(rd.get('don_gia') or 0)).number_format = '#,##0'
            ws.cell(cur_row, 13, float(rd.get('thanh_tien') or 0)).number_format = '#,##0'
            ws.cell(cur_row, 14, rd.get('ghi_chu') or '')
            ws.cell(cur_row, 15, rd.get('ket_qua_doi_soat') or 'OK')

            for c in range(1, 16):
                ws.cell(cur_row, c).font = Font(name='Arial', size=9)
                ws.cell(cur_row, c).border = thin_border
            stt += 1
            cur_row += 1
        end_dau = cur_row - 1

        # Sum Dau
        ws.cell(cur_row, 1, 'Tổng cộng máy chạy dầu')
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=9)
        ws.cell(cur_row, 10, f'=SUM(J{start_dau}:J{end_dau})').number_format = '0.00'
        ws.cell(cur_row, 11, f'=SUM(K{start_dau}:K{end_dau})').number_format = '0.00'
        ws.cell(cur_row, 13, f'=SUM(M{start_dau}:M{end_dau})').number_format = '#,##0'
        for c in range(1, 16):
            ws.cell(cur_row, c).font = Font(name='Arial', size=9, bold=True, color='B45309')
            ws.cell(cur_row, c).fill = light_amber
            ws.cell(cur_row, c).border = thick_bottom
        sum_dau_row = cur_row
        cur_row += 1

        # Grand total
        ws.cell(cur_row, 1, 'TỔNG CỘNG (XĂNG + DẦU)')
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=9)
        ws.cell(cur_row, 10, f'=J{sum_xang_row}+J{sum_dau_row}').number_format = '0.00'
        ws.cell(cur_row, 11, f'=K{sum_xang_row}+K{sum_dau_row}').number_format = '0.00'
        ws.cell(cur_row, 13, f'=M{sum_xang_row}+M{sum_dau_row}').number_format = '#,##0'
        for c in range(1, 16):
            ws.cell(cur_row, c).font = Font(name='Arial', size=10, bold=True, color='000000')
            ws.cell(cur_row, c).fill = sum_fill
            ws.cell(cur_row, c).border = double_bottom

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['M'].width = 20

    def add_hd_sheet(sheet_title, target_invoices, group_name):
        ws = wb.create_sheet(title=sheet_title)
        ws['A1'] = 'TRUNG TÂM MẠNG LƯỚI MOBIFONE MIỀN NAM'; ws['A1'].font = Font(name='Arial', size=10, bold=True)
        ws['A2'] = 'ĐƠN VỊ: ĐÀI VIỄN THÔNG ĐỒNG NAI'; ws['A2'].font = Font(name='Arial', size=10, bold=True)
        ws['A4'] = f'BẢNG KÊ HÓA ĐƠN NHIÊN LIỆU MÁY PHÁT ĐIỆN THÁNG {month:02d}/{year}'
        ws['A4'].font = Font(name='Arial', size=13, bold=True, color='047857')
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A4:O4')
        ws['A5'] = f'({group_name})'; ws['A5'].font = Font(name='Arial', size=10, italic=True)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A5:O5')

        headers_hd = [
            'STT', 'Ngày lập chứng từ', 'Số chứng từ (HĐ)', 'Đơn vị xuất hóa đơn',
            'Loại NL', 'Diễn giải', 'Số lượng (Lít)', 'Mã số thuế', 'Mẫu số HĐ',
            'Ký hiệu HĐ', 'Thành tiền trước VAT', 'Thuế VAT', 'Tổng cộng thanh toán',
            'Link tra cứu hóa đơn', 'Mã tra cứu'
        ]
        for c, h in enumerate(headers_hd, 1):
            cell = ws.cell(7, c, h)
            cell.fill = header_fill_green
            cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        cur_row = 8
        stt = 1
        for inv in target_invoices:
            items = inv.get('items') or []
            if isinstance(items, str):
                import json
                try: items = json.loads(items)
                except: items = []
            
            is_xang = False
            qty = 0
            dien_giai = "Nhiên liệu chạy máy"
            for it in items:
                name = (it.get('ten') or it.get('name') or '').lower()
                q = float(it.get('sl') or it.get('quantity') or 0)
                qty += q
                if 'xăng' in name or 'xang' in name or 'ron' in name:
                    is_xang = True
                    dien_giai = "Xăng RON 95"
                elif 'dầu' in name or 'dau' in name or 'diesel' in name:
                    dien_giai = "Dầu Điêzen"

            sub = float(inv.get('sub_total') or 0)
            vat = float(inv.get('vat_amount') or 0)
            tot = float(inv.get('total_amount') or (sub + vat))

            ws.cell(cur_row, 1, f"L{stt}")
            ws.cell(cur_row, 2, inv.get('invoice_date') or '')
            ws.cell(cur_row, 3, inv.get('invoice_number') or '')
            ws.cell(cur_row, 4, inv.get('seller_name') or '')
            ws.cell(cur_row, 5, 'Xăng' if is_xang else 'Dầu')
            ws.cell(cur_row, 6, dien_giai)
            ws.cell(cur_row, 7, qty).number_format = '0.0'
            ws.cell(cur_row, 8, inv.get('seller_mst') or '')
            ws.cell(cur_row, 9, inv.get('mau_so') or '')
            ws.cell(cur_row, 10, inv.get('kh_hd') or '')
            ws.cell(cur_row, 11, sub).number_format = '#,##0'
            ws.cell(cur_row, 12, vat).number_format = '#,##0'
            ws.cell(cur_row, 13, tot).number_format = '#,##0'
            ws.cell(cur_row, 14, inv.get('invoice_url') or '')
            ws.cell(cur_row, 15, inv.get('ma_tra_cuu') or '')

            for c in range(1, 16):
                ws.cell(cur_row, c).font = Font(name='Arial', size=9)
                ws.cell(cur_row, c).border = thin_border
            stt += 1
            cur_row += 1

        end_inv = cur_row - 1
        ws.cell(cur_row, 1, 'TỔNG CỘNG TOÀN BỘ HÓA ĐƠN')
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
        ws.cell(cur_row, 7, f'=SUM(G8:G{end_inv})').number_format = '0.0'
        ws.cell(cur_row, 11, f'=SUM(K8:K{end_inv})').number_format = '#,##0'
        ws.cell(cur_row, 12, f'=SUM(L8:L{end_inv})').number_format = '#,##0'
        ws.cell(cur_row, 13, f'=SUM(M8:M{end_inv})').number_format = '#,##0'
        for c in range(1, 16):
            ws.cell(cur_row, c).font = Font(name='Arial', size=10, bold=True, color='047857')
            ws.cell(cur_row, c).fill = sum_fill
            ws.cell(cur_row, c).border = double_bottom

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['D'].width = 38
        ws.column_dimensions['N'].width = 36
        ws.column_dimensions['O'].width = 24

    if month >= 8 and year >= 2026:
        # Split 2 Groups
        g1_logs = [l for l in logs if is_group_1(l.get('site_id'), sites.get(l.get('site_id'), {}).get('site_id_old'))]
        g2_logs = [l for l in logs if not is_group_1(l.get('site_id'), sites.get(l.get('site_id'), {}).get('site_id_old'))]

        g1_invs = [i for i in invoices if '0100686209-129' in (i.get('buyer_mst') or '') or 'ĐỒNG NAI' in (i.get('buyer_name') or '').upper()]
        g2_invs = [i for i in invoices if not ('0100686209-129' in (i.get('buyer_mst') or '') or 'ĐỒNG NAI' in (i.get('buyer_name') or '').upper())]

        add_02a_sheet('02A_TTNB_DongNai_67Tram', g1_logs, 'MobiFone Đồng Nai - 67 Trạm Đặc Thù')
        add_hd_sheet('HD_DongNai_67Tram', g1_invs, 'MobiFone Đồng Nai - 67 Trạm Đặc Thù')
        add_02a_sheet('02A_TTNB_ToanCau', g2_logs, 'MobiFone Toàn Cầu')
        add_hd_sheet('HD_ToanCau', g2_invs, 'MobiFone Toàn Cầu')
    else:
        add_02a_sheet('02A-TTNB_NLMPD', logs, 'Toàn bộ trạm Đài Viễn thông Đồng Nai')
        add_hd_sheet('HD', invoices, 'Toàn bộ hóa đơn nhiên liệu')

    if not output_path:
        output_path = f"/Users/cang_it/Desktop/Ho_So_Thanh_Toan_Chuan_Mau_{month:02d}_{year}.xlsx"

    wb.save(output_path)
    print(f"✅ Đã xuất thành công hồ sơ chuẩn mẫu ra: {output_path}")
    return output_path

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--month', type=int, default=8, help='Tháng đối soát (1-12)')
    parser.add_argument('--year', type=int, default=2026, help='Năm đối soát')
    parser.add_argument('--out', type=str, default='', help='Đường dẫn file đầu ra')
    args = parser.parse_args()

    out_file = args.out if args.out else None
    create_styled_workbook(month=args.month, year=args.year, output_path=out_file)
